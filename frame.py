from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("DATA ENTRY FORM")
root.geometry('900x900+300+200')
root.resizable(False,False)
root.configure(bg="#326273")



file=pathlib.Path('abc.xlxs')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="Contact No"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Adress"

    file.save('abc.xlxs')
    
def Submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    adress=adressEntry.get(1.0,END)

    file=openpyxl.load_workbook('abc.xlxs')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=adress)

    file.save(r'abc.xlxs')

    messagebox.showinfo('info','details added!')
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    adressEntry.delete(1.0,END)

def Clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    adressEntry.delete(1.0,END)

#heading
Label(root,text="REGISTRATION FORM",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#label
Label(root,text='NAME:',font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text='CONTACT NO:',font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text='AGE:',font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text='GENDER:',font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text='ADRESS:',font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20).place(x=200,y=100)
contactEntry = Entry(root,textvariable=contactValue,width=45,bd=2,font=20).place(x=200,y=150)
ageEntry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20).place(x=200,y=200)

#gender
gender_combobox = Combobox(root,values=['Male','Female','Transgender'],font='arial 14',state='r',width=14).place(x=470,y=200)

adressEntry =Text(root,width=50,bd=4,height=4).place(x=200,y=250)

Button(root,text="Submit",bg="#326273",fg="White",width=15,height=2,command=Submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="White",width=15,height=2,command=Clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg="White",width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)

root.mainloop()
